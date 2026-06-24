"""数据导入页面

本模块实现抽卡分析器的数据导入功能界面，支持三种导入方式：
1. 自动获取 - 自动从游戏缓存中提取API URL并获取抽卡记录
2. 手动粘贴URL - 用户通过抓包获取URL后粘贴到输入框
3. 文件导入 - 从JSON/Excel/CSV文件导入抽卡记录

支持的游戏：
    - 原神(genshin)
    - 星穹铁道(starrail)
    - 绝区零(zzz)
    - 鸣潮(wutheringwaves)
    - 终末地(endfield)
    - 明日方舟(arknights)

Qt信号/槽架构：
    FetchThread (QThread) 与 ImportWidget 的通信：
        - progress(str, float) → _on_progress：更新进度文字和进度条
        - finished(list)       → _on_fetch_done / _on_game_fetch_done：获取完成
        - error(str)           → _on_fetch_error / _on_game_fetch_error：获取失败

    线程安全说明：
        - QThread的信号发射是线程安全的（跨线程排队调用）
        - 通过信号将数据传回主线程处理，避免跨线程UI操作
        - cancel() 通过标志位实现协作式取消，fetcher在每轮循环中检查

异常处理：
    - 网络超时、API错误、authkey过期等都有对应的错误提示
    - 文件解析异常会弹窗显示具体错误信息
    - 获取线程异常通过 error 信号传回主线程处理
"""

# ========== 标准库导入 ==========

import json
# json: JSON数据解析和序列化库，用于解析导入的JSON格式抽卡记录

import os
# os: 操作系统接口，提供文件路径操作等功能（本文件中未直接使用，可能为其他模块预留）

import csv
# csv: CSV文件读写库，用于解析CSV格式的抽卡记录文件

# ========== PySide6/Qt导入 ==========

from PySide6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QLabel, QFrame,
    QPushButton, QLineEdit, QTextEdit, QFileDialog, QMessageBox,
    QProgressBar, QGroupBox, QGridLayout, QComboBox, QDialog
)
# QWidget: 控件基类
# QVBoxLayout: 垂直布局
# QHBoxLayout: 水平布局
# QLabel: 文本标签
# QFrame: 带边框容器
# QPushButton: 按钮
# QLineEdit: 单行文本输入框
# QTextEdit: 多行文本编辑框（用于日志输出）
# QFileDialog: 文件选择对话框
# QMessageBox: 消息提示框（信息/警告/错误/确认）
# QProgressBar: 进度条
# QGroupBox: 分组框（带标题的容器）
# QGridLayout: 网格布局
# QComboBox: 下拉选择框
# QDialog: 对话框基类

from PySide6.QtCore import Qt, QThread, Signal, QTimer
# Qt: Qt核心常量
# QThread: 线程基类，用于在后台执行耗时操作（如网络请求）
# Signal: 信号类，用于线程间通信（跨线程排队调用槽函数）
# QTimer: 定时器，用于延迟执行操作（如登录成功后延迟调用URL获取）

from PySide6.QtGui import QFont
# QFont: 字体类

# ========== 业务模块导入 ==========

from core.database import Database
# Database: 数据库访问层，封装了SQLite的增删改查操作

from core.models import Account, GachaRecord, GAME_NAMES
# Account: 账号数据模型，包含id, game, uid, nickname, server等字段
# GachaRecord: 抽卡记录数据模型，包含account_id, game, pool_type, item_name, rarity等字段
# GAME_NAMES: 游戏ID到中文名称的映射字典，如{"genshin": "原神", ...}

from fetchers import get_fetcher
# get_fetcher(game): 根据游戏ID返回对应的fetcher实例（如MihoyoAPI、HypergryphAPI等）

from fetchers.url_parser import URLParser
# URLParser: URL解析器，从URL中提取游戏类型、region、uid等信息


class FetchThread(QThread):
    """后台获取抽卡记录的线程

    继承QThread，在run()方法中执行网络请求。
    通过信号（Signal）与主线程通信，避免跨线程UI操作。

    信号定义：
        progress(str, float): 进度更新信号
            - str: 进度描述文字（如"正在获取第3页..."）
            - float: 进度值（0.0-1.0之间，-1表示不确定进度）
        finished(list): 获取完成信号，传递记录列表
        error(str): 获取失败信号，传递错误描述文字

    线程生命周期：
        1. 创建FetchThread实例
        2. 连接信号到槽函数
        3. 调用start()启动线程
        4. 线程执行run()中的任务
        5. 完成后发射finished或error信号
        6. 可通过cancel()请求协作式取消

    取消机制：
        - cancel()设置 _cancelled 标志位为True
        - fetcher在每轮循环中检查 _cancel_check 回调
        - 如果检测到取消，fetcher会提前终止并返回已获取的记录
        - run()中也会检查 _cancelled 并发射error信号
    """

    # ===== 信号定义 =====
    # progress信号：(消息文字, 进度值0.0-1.0)
    progress = Signal(str, float)
    # finished信号：(记录列表)
    finished = Signal(list)
    # error信号：(错误消息)
    error = Signal(str)

    def __init__(self, game, url=None, account_id=None, latest_time=None):
        """构造函数

        参数：
            game (str): 游戏ID，如"genshin"、"starrail"、"endfield"
            url (str|None): 抽卡记录API的完整URL，None表示自动检测
            account_id (int|None): 账号ID，用于标记记录所属账号
            latest_time (str|None): 已有记录的最新时间，用于增量获取
                                   传入后只获取该时间之后的新记录
                                   None表示获取全部记录

        实例属性：
            _cancelled (bool): 取消标志位，True表示已请求取消
            detected_uid (str): fetcher检测到的UID，获取完成后可读取
            _fetcher_instance: fetcher实例引用，用于取消时终止子进程
        """
        # 调用QThread构造函数
        super().__init__()
        # 游戏标识符
        self.game = game
        # 抽卡记录URL（可选，为None时由fetcher自动检测）
        self.url = url
        # 账号ID（用于将记录绑定到特定账号）
        self.account_id = account_id
        # 增量获取的起始时间（已有记录的最新时间）
        self.latest_time = latest_time
        # 取消标志位，初始为False（未取消）
        self._cancelled = False
        # fetcher检测到的UID（获取完成后可读取），初始为空字符串
        self.detected_uid = ""

    def cancel(self):
        """请求取消获取

        设置取消标志位，fetcher在下次循环检查时会检测到并终止。
        这是一种协作式取消机制，不会强制中断正在进行的网络请求。
        """
        self._cancelled = True

    def is_cancelled(self):
        """检查是否已请求取消

        返回：
            bool: True表示已请求取消，False表示未取消

        此方法被设置为fetcher的 _cancel_check 回调，
        fetcher在每轮循环中调用此方法检查是否需要终止。
        """
        return self._cancelled

    def run(self):
        """线程执行入口（QThread的虚函数重写）

        在新线程中执行以下操作：
            1. 通过 get_fetcher() 获取对应游戏的fetcher实例
            2. 设置进度回调和取消检查回调
            3. 调用 fetcher.fetch_records() 执行网络请求
            4. 根据结果发射 finished 或 error 信号

        异常处理：
            - 如果已取消：发射error("用户已取消获取")
            - 其他异常：发射error(str(e))

        注意：
            - 此方法在子线程中执行，不能直接操作UI
            - 通过信号将结果传回主线程处理
        """
        try:
            # 获取对应游戏的fetcher实例
            # 不同游戏使用不同的API接口和请求格式
            fetcher = get_fetcher(self.game)
            # 保存fetcher实例引用，取消时可能需要终止其子进程
            self._fetcher_instance = fetcher
            # 设置进度回调：fetcher内部调用此函数更新进度
            # lambda将fetcher的回调参数转发到QThread的progress信号
            fetcher.set_progress_callback(lambda msg, p: self.progress.emit(msg, p))
            # 设置取消检查回调：fetcher在循环中调用此函数判断是否需要终止
            fetcher._cancel_check = self.is_cancelled
            # 执行网络请求获取抽卡记录
            records = fetcher.fetch_records(url=self.url, account_id=self.account_id, latest_time=self.latest_time)
            # 从fetcher获取检测到的UID（某些fetcher会在请求过程中检测UID）
            self.detected_uid = getattr(fetcher, '_detected_uid', '')
            # 再次检查是否已取消
            if self._cancelled:
                # 用户在获取过程中请求了取消
                self.error.emit("用户已取消获取")
            else:
                # 正常完成，发射完成信号
                self.finished.emit(records)
        except Exception as e:
            # 捕获所有异常
            if self._cancelled:
                # 取消导致的异常也当作取消处理
                self.error.emit("用户已取消获取")
            else:
                # 其他异常，将异常信息传回主线程
                self.error.emit(str(e))


class ImportWidget(QWidget):
    """数据导入页面主控件

    提供三种数据导入方式的UI和逻辑：
    1. 自动获取 - 从游戏缓存读取URL并自动获取记录（支持批量多游戏）
    2. 手动粘贴URL - 从抓包工具获取URL后解析获取
    3. 文件导入 - 从JSON/Excel/CSV文件导入

    对于终末地和明日方舟，还支持登录获取（通过鹰角账号登录获取Token）。

    信号/槽连接汇总：
        - auto_fetch_btn.clicked → _auto_fetch：自动获取按钮
        - url_fetch_btn.clicked → _url_fetch：URL获取按钮
        - url_input (回车键) → _url_fetch：URL输入框回车触发
        - paste_btn.clicked → _paste_from_clipboard：粘贴按钮
        - login_btn.clicked → _login_fetch：登录获取按钮
        - auto_game_combo.currentIndexChanged → _update_login_btn_visibility
        - cancel_btn.clicked → _cancel_fetch：取消获取按钮

    数据流：
        1. 自动获取：CacheReader.extract_url() → FetchThread → _on_game_fetch_done()
        2. URL获取：URLParser.parse() → FetchThread → _on_fetch_done()
        3. 文件导入：QFileDialog → _parse_file() → db.add_records()

    属性说明：
        main_window: 主窗口引用
        db: 数据库实例
        fetch_thread: FetchThread实例引用（获取进行中时非None）
        _detected_games: 自动获取时检测到的游戏列表 [(game_id, url), ...]
        _current_fetch_index: 当前正在获取的游戏索引（批量获取时使用）
    """

    def __init__(self, main_window):
        """构造函数

        参数：
            main_window: 主窗口实例，提供 get_current_game()、set_account()、
                        refresh_all() 等方法
        """
        super().__init__()
        # 保存主窗口引用
        self.main_window = main_window
        # 初始化数据库访问对象
        self.db = Database()
        # 后台获取线程引用，初始为None（无正在进行的获取任务）
        self.fetch_thread = None
        # 调用UI初始化方法
        self._init_ui()

    def _init_ui(self):
        """初始化界面布局

        页面结构：
        QVBoxLayout
        ├── QLabel "获取数据" (标题)
        ├── QLabel "选择一种方式导入抽卡记录" (副标题)
        ├── QGroupBox "方式一：自动获取（推荐）"
        │   ├── QHBoxLayout (游戏选择 + 开始获取按钮)
        │   └── QLabel (使用说明)
        ├── QGroupBox "方式二：手动粘贴URL"
        │   ├── QLabel (使用说明)
        │   ├── QLineEdit (URL输入框)
        │   └── QHBoxLayout (解析获取 + 粘贴 + 登录获取按钮)
        ├── QGroupBox "方式三：文件导入"
        │   ├── QLabel (支持格式说明)
        │   └── QHBoxLayout (JSON + Excel + CSV 按钮)
        ├── QProgressBar (进度条，初始隐藏)
        ├── QHBoxLayout (状态文字 + 取消按钮)
        ├── QLabel "操作日志" (标题)
        ├── QTextEdit (日志输出，只读)
        └── addStretch() 底部弹性空间
        """
        # 创建主垂直布局
        layout = QVBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 0)

        # ===== 标题区域 =====
        title = QLabel("获取数据")
        # 设置标题字体：微软雅黑，18号，加粗
        title.setFont(QFont("Microsoft YaHei", 18, QFont.Weight.Bold))
        layout.addWidget(title)

        # 副标题
        subtitle = QLabel("选择一种方式导入抽卡记录")
        subtitle.setStyleSheet("color: #888; font-size: 13px;")
        layout.addWidget(subtitle)

        # ===== 方式一：自动获取 =====
        # QGroupBox 创建带标题的分组框，视觉上将相关控件归组
        auto_group = QGroupBox("方式一：自动获取（推荐）")
        auto_group.setStyleSheet("QGroupBox { font-weight: bold; font-size: 14px; }")
        auto_layout = QVBoxLayout(auto_group)

        # 游戏选择行
        select_layout = QHBoxLayout()
        select_layout.addWidget(QLabel("游戏:"))
        # 游戏选择下拉框
        self.auto_game_combo = QComboBox()
        # 定义游戏选项列表：(游戏ID, 显示名称)
        self._auto_game_options = [
            ("all", "全部游戏"),
            ("genshin", "原神"),
            ("starrail", "星穹铁道"),
            ("zzz", "绝区零"),
            ("wutheringwaves", "鸣潮"),
            ("endfield", "终末地"),
            ("arknights", "明日方舟"),
        ]
        # 遍历选项列表，添加到下拉框
        for gid, gname in self._auto_game_options:
            # addItem(text, data): data参数关联到选项，通过currentData()获取
            self.auto_game_combo.addItem(gname, gid)
        select_layout.addWidget(self.auto_game_combo)

        # "开始获取"按钮
        self.auto_fetch_btn = QPushButton("开始获取")
        self.auto_fetch_btn.setFixedSize(100, 32)
        # 连接点击信号到 _auto_fetch 槽函数
        self.auto_fetch_btn.clicked.connect(self._auto_fetch)
        select_layout.addWidget(self.auto_fetch_btn)
        # 右侧弹性空间
        select_layout.addStretch()
        auto_layout.addLayout(select_layout)

        # 使用说明文字
        auto_desc = QLabel("请确保已打开对应游戏并进入抽卡记录页面")
        auto_desc.setStyleSheet("color: #666;")
        auto_layout.addWidget(auto_desc)

        # 将方式一的分组框添加到主布局
        layout.addWidget(auto_group)

        # ===== 方式二：手动粘贴URL =====
        url_group = QGroupBox("方式二：手动粘贴URL")
        url_group.setStyleSheet("QGroupBox { font-weight: bold; font-size: 14px; }")
        url_layout = QVBoxLayout(url_group)

        # URL获取方式的使用说明
        url_desc = QLabel(
            "通过抓包工具获取抽卡记录API的完整URL，粘贴到下方输入框。\n"
            "终末地：粘贴鹰角账号Token（从 https://user.hypergryph.com/ 登录获取）"
        )
        url_desc.setStyleSheet("color: #666;")
        url_layout.addWidget(url_desc)

        # URL输入框
        self.url_input = QLineEdit()
        # 设置占位提示文字（输入框为空时显示的灰色提示）
        self.url_input.setPlaceholderText("粘贴抽卡记录URL 或 终末地账号Token...")
        url_layout.addWidget(self.url_input)

        # 按钮行
        url_btn_layout = QHBoxLayout()
        # "解析获取"按钮
        self.url_fetch_btn = QPushButton("解析获取")
        self.url_fetch_btn.setFixedSize(120, 36)
        self.url_fetch_btn.clicked.connect(self._url_fetch)
        url_btn_layout.addWidget(self.url_fetch_btn)

        # "从剪贴板粘贴"按钮
        paste_btn = QPushButton("从剪贴板粘贴")
        paste_btn.setFixedSize(120, 36)
        # 灰色背景，视觉上次要操作
        paste_btn.setStyleSheet("background-color: #666;")
        paste_btn.clicked.connect(self._paste_from_clipboard)
        url_btn_layout.addWidget(paste_btn)

        # "登录获取"按钮（仅终末地和明日方舟显示）
        self.login_btn = QPushButton("登录获取")
        self.login_btn.setFixedSize(120, 36)
        # 橙色背景，突出显示
        self.login_btn.setStyleSheet("background-color: #E65100;")
        self.login_btn.setToolTip("登录鹰角账号获取Token（仅终末地和明日方舟）")
        self.login_btn.clicked.connect(self._login_fetch)
        url_btn_layout.addWidget(self.login_btn)

        # 连接游戏选择下拉框信号，控制登录按钮显示
        # currentIndexChanged 在下拉框选中项变化时发射
        self.auto_game_combo.currentIndexChanged.connect(self._update_login_btn_visibility)
        # 初始化登录按钮显示状态
        self._update_login_btn_visibility()

        # 右侧弹性空间
        url_btn_layout.addStretch()
        url_layout.addLayout(url_btn_layout)

        # 将方式二的分组框添加到主布局
        layout.addWidget(url_group)

        # ===== 方式三：文件导入 =====
        file_group = QGroupBox("方式三：文件导入")
        file_group.setStyleSheet("QGroupBox { font-weight: bold; font-size: 14px; }")
        file_layout = QVBoxLayout(file_group)

        # 支持的文件格式说明
        file_desc = QLabel("支持 JSON、Excel (.xlsx)、CSV 格式")
        file_desc.setStyleSheet("color: #666;")
        file_layout.addWidget(file_desc)

        # 文件导入按钮行
        file_btn_layout = QHBoxLayout()
        # JSON导入按钮
        json_btn = QPushButton("导入 JSON")
        # lambda表达式将file_type参数传递给 _import_file
        json_btn.clicked.connect(lambda: self._import_file("json"))
        # Excel导入按钮
        excel_btn = QPushButton("导入 Excel")
        excel_btn.clicked.connect(lambda: self._import_file("excel"))
        # CSV导入按钮
        csv_btn = QPushButton("导入 CSV")
        csv_btn.clicked.connect(lambda: self._import_file("csv"))

        file_btn_layout.addWidget(json_btn)
        file_btn_layout.addWidget(excel_btn)
        file_btn_layout.addWidget(csv_btn)
        # 右侧弹性空间
        file_btn_layout.addStretch()
        file_layout.addLayout(file_btn_layout)

        # 将方式三的分组框添加到主布局
        layout.addWidget(file_group)

        # ===== 进度条 =====
        self.progress_bar = QProgressBar()
        # 初始隐藏，获取开始时显示
        self.progress_bar.setVisible(False)
        layout.addWidget(self.progress_bar)

        # ===== 状态栏 =====
        status_layout = QHBoxLayout()
        # 状态文字标签
        self.status_label = QLabel("")
        self.status_label.setStyleSheet("color: #666;")
        status_layout.addWidget(self.status_label)
        # 左侧弹性空间（使取消按钮靠右）
        status_layout.addStretch()

        # "取消获取"按钮
        self.cancel_btn = QPushButton("取消获取")
        self.cancel_btn.setFixedSize(80, 28)
        # 初始隐藏
        self.cancel_btn.setVisible(False)
        # 红色背景，白色文字，无边框，圆角
        self.cancel_btn.setStyleSheet("background-color: #F44336; color: white; border: none; border-radius: 4px;")
        self.cancel_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        self.cancel_btn.clicked.connect(self._cancel_fetch)
        status_layout.addWidget(self.cancel_btn)

        # 将状态栏添加到主布局
        layout.addLayout(status_layout)

        # ===== 操作日志 =====
        log_label = QLabel("操作日志")
        log_label.setFont(QFont("Microsoft YaHei", 12, QFont.Weight.Bold))
        layout.addWidget(log_label)

        # 日志输出文本框
        self.log_text = QTextEdit()
        # 设置为只读，用户不能编辑日志内容
        self.log_text.setReadOnly(True)
        # 限制最大高度150像素
        self.log_text.setMaximumHeight(150)
        # 设置样式：白色背景、黑色文字、Consolas等宽字体（便于对齐）
        self.log_text.setStyleSheet(
            "background-color: #ffffff; color: #000000; font-family: Consolas; font-size: 12px;"
        )
        layout.addWidget(self.log_text)

        # 底部弹性空间
        layout.addStretch()

    def _log(self, message):
        """向日志区域追加一条消息

        参数：
            message (str): 要追加的日志文字，支持富文本格式
        """
        # QTextEdit.append() 会在末尾追加文字并自动滚动到最新位置
        self.log_text.append(message)

    def _auto_fetch(self):
        """自动获取 - 自动检测选中的游戏

        流程：
            1. 读取下拉框选中的游戏
            2. 检查是否有不支持自动获取的游戏（终末地/明日方舟）
            3. 对支持的游戏，使用CacheReader从游戏缓存中提取API URL
            4. 依次为每个检测到的游戏启动获取线程

        不支持自动获取的游戏会弹窗提示用户使用登录获取。
        如果混合了支持和不支持的游戏，跳过不支持的继续获取支持的游戏。
        """
        # 延迟导入，避免循环引用（这些模块可能引用本模块）
        from fetchers.cache_reader import CacheReader
        from fetchers.url_parser import URLParser

        # 读取下拉框选中的游戏ID
        selected_id = self.auto_game_combo.currentData()
        if selected_id == "all":
            # "全部游戏"：选中所有支持的游戏
            selected = ["genshin", "starrail", "zzz", "wutheringwaves", "endfield", "arknights"]
        else:
            selected = [selected_id]

        # ===== 检查不支持自动获取的游戏 =====
        unsupported_games = []  # 不支持的游戏中文名列表
        supported_games = []    # 支持的游戏ID列表
        for game_id in selected:
            if game_id in ["endfield", "arknights"]:
                # 终末地和明日方舟不支持缓存读取，需要通过Token获取
                unsupported_games.append(GAME_NAMES.get(game_id, game_id))
            else:
                supported_games.append(game_id)

        # 弹窗提示不支持的游戏
        if unsupported_games:
            game_names = "、".join(unsupported_games)
            QMessageBox.warning(
                self, "提示",
                f"{game_names}暂时不支持自动获取，\n请使用方法二登录获取。"
            )
            # 如果全部都不支持，直接返回
            if not supported_games:
                return
            # 混合情况：只继续获取支持的游戏
            selected = supported_games

        # 设置UI为"获取中"状态（禁用按钮、显示进度条）
        self._set_fetching(True)
        # 记录日志
        game_label = self.auto_game_combo.currentText()
        self._log(f"开始自动检测: {game_label}...")

        # 初始化缓存读取器
        cache = CacheReader()
        detected_games = []  # 检测到的游戏列表 [(game_id, url), ...]

        # 逐个游戏扫描缓存提取URL
        for game_id in selected:
            try:
                # 从游戏缓存文件中提取抽卡记录API的URL
                url = cache.extract_url(game_id)
                if url:
                    detected_games.append((game_id, url))
            except Exception as e:
                # 扫描失败不中断，记录日志继续下一个
                self._log(f"  ✗ 扫描 {GAME_NAMES.get(game_id, game_id)} 失败: {str(e)}")

        # 未找到任何游戏记录
        if not detected_games:
            self._log("\n未找到任何游戏记录！")
            self._log("请确保：")
            self._log("1. 已打开游戏")
            self._log("2. 进入抽卡/跃迁记录页面")
            self._log("3. 等待记录加载完成")
            self._log("4. 切回本程序重试")
            # 恢复UI状态
            self._set_fetching(False)
            QMessageBox.information(self, "提示",
                "未找到任何游戏记录。\n\n"
                "请确保：\n"
                "1. 已打开游戏\n"
                "2. 进入抽卡/跃迁记录页面\n"
                "3. 等待记录加载完成\n"
                "4. 切回本程序重试\n\n"
                "如果还是找不到，请尝试手动粘贴URL。")
            return

        self._log(f"\n共检测到 {len(detected_games)} 个游戏，开始获取记录...")

        # 保存检测到的游戏列表，开始依次获取
        self._detected_games = detected_games
        # 当前获取索引（从0开始）
        self._current_fetch_index = 0
        # 开始获取第一个游戏
        self._fetch_next_game()

    def _fetch_next_game(self):
        """获取下一个游戏的记录（批量获取的迭代器）

        检查是否还有待获取的游戏，如果有则启动FetchThread获取。
        获取完成后会调用自身（递归式迭代），直到所有游戏处理完毕。

        流程：
            1. 检查是否还有待处理的游戏
            2. 如果全部完成：恢复UI、刷新主窗口、弹窗提示
            3. 如果还有：自动检测/创建账号 → 获取最新时间 → 启动FetchThread
        """
        # 检查是否所有游戏都已处理完毕
        if self._current_fetch_index >= len(self._detected_games):
            # 所有游戏获取完成
            self._set_fetching(False)
            self._log("\n所有游戏获取完成！")
            # 刷新主窗口的所有页面数据
            self.main_window.refresh_all()
            QMessageBox.information(self, "完成", "所有游戏记录获取完成！")
            return

        # 获取当前待处理的游戏
        game, url = self._detected_games[self._current_fetch_index]
        self._log(f"\n正在获取 {GAME_NAMES.get(game, game)}...")

        # 自动检测或创建账号
        account = self._auto_detect_account(game, url)
        if not account:
            # 账号检测失败（如用户取消输入UID），跳过该游戏
            self._log(f"  跳过 {GAME_NAMES.get(game, game)}")
            self._current_fetch_index += 1
            # 递归获取下一个游戏
            self._fetch_next_game()
            return

        # 切换主窗口到当前游戏和账号
        self.main_window._on_game_changed(game)
        self.main_window.set_account(account)

        # 获取最新记录时间，用于增量获取
        latest_time = None
        records = self.db.get_records(account.id)
        if records:
            # 取所有记录中最新的时间
            latest_time = max(r.time for r in records if r.time)
            self._log(f"  已有记录，从 {latest_time} 开始增量获取")

        # 创建后台获取线程
        self.fetch_thread = FetchThread(game, url=url, account_id=account.id, latest_time=latest_time)
        # 连接信号到对应的槽函数
        self.fetch_thread.progress.connect(self._on_progress)
        # 注意：批量获取完成后的处理用 _on_game_fetch_done（会继续获取下一个游戏）
        self.fetch_thread.finished.connect(self._on_game_fetch_done)
        self.fetch_thread.error.connect(self._on_game_fetch_error)
        # 启动线程（异步执行，不阻塞UI）
        self.fetch_thread.start()

    def _generate_nickname(self, game, uid):
        """生成唯一的游戏昵称

        格式：{游戏称呼}{UID后N位}
        例如原神UID为123456789 → "旅行者789"

        如果生成的昵称已存在（被其他账号使用），会逐步增加UID的位数，
        从后3位增加到后4位、5位...最多8位。如果仍然冲突，添加序号后缀。

        参数：
            game (str): 游戏ID
            uid (str): 玩家UID

        返回：
            str: 唯一的昵称字符串

        数据结构：
            game_titles: 游戏ID到称呼的映射
            existing: 当前游戏已有昵称的集合（用于查重）
        """
        # 游戏到称呼的映射
        game_titles = {
            "genshin": "旅行者", "starrail": "开拓者", "zzz": "绳匠",
            "wutheringwaves": "漂泊者", "endfield": "管理员", "arknights": "博士",
        }
        # 获取对应游戏的称呼
        title = game_titles.get(game, "玩家")
        # 收集当前游戏所有已有账号的昵称（用于查重）
        existing = {a.nickname for a in self.db.get_accounts(game)}

        # 从后3位UID开始，逐步增加位数直到找到不重复的昵称
        for digits in range(3, min(len(uid), 8) + 1):
            nickname = f"{title}{uid[-digits:]}"
            if nickname not in existing:
                return nickname

        # 极端情况：所有位数都冲突，加序号后缀
        # 如 "旅行者789(2)", "旅行者789(3)", ...
        for i in range(2, 100):
            nickname = f"{title}{uid[-3:]}({i})"
            if nickname not in existing:
                return nickname

        # 理论上不会到这里（100个序号应该够用）
        return f"{title}{uid[-3:]}"

    def _auto_detect_account(self, game, url, detected_uid=None):
        """自动检测并创建账号

        根据游戏和URL自动检测已有账号或创建新账号。
        支持多种检测方式：
            1. 从URL参数中提取UID
            2. 从fetcher预先检测到的UID（detected_uid参数）
            3. 从游戏缓存文件中读取UID（UidInfo.txt）
            4. 从缓存中读取昵称
            5. 对于明日方舟/终末地：如果没有UID，让用户手动输入

        参数：
            game (str): 游戏ID
            url (str): 抽卡记录URL或Token
            detected_uid (str|None): fetcher预先检测到的UID（可选）

        返回：
            Account|None: 检测或创建的账号对象，None表示用户取消

        检测优先级：
            1. 数据库中已有的相同UID的账号（直接返回）
            2. URL中的UID + 缓存中的昵称 → 创建新账号
            3. 明日方舟/终末地无UID → 用户手动输入 → 创建新账号
            4. 复用该游戏的第一个已有账号
        """
        from fetchers.url_parser import URLParser
        from fetchers.cache_reader import CacheReader
        from PySide6.QtWidgets import QInputDialog

        # 解析URL，提取game、region、params等信息
        parsed = URLParser.parse(url)
        region = parsed.get("region", "cn")

        # 尝试从URL参数中提取UID
        uid = None
        params = parsed.get("params", {})
        # 尝试多种UID参数名（不同游戏可能使用不同的参数名）
        uid = params.get("uid") or params.get("user_id") or params.get("player_id")

        # 使用fetcher预先检测到的UID（明日方舟/终末地可能在此前就已检测到）
        if not uid and detected_uid:
            uid = detected_uid

        # 如果URL中没有UID，尝试从游戏缓存文件中读取
        if not uid:
            cache = CacheReader()
            uid = cache.extract_uid(game, region)
            if uid:
                self._log(f"  从游戏文件读取到UID: {uid}")

        # 尝试从缓存中获取玩家昵称
        nickname = ""
        cache = CacheReader()
        nickname = cache.extract_nickname(game, region)
        if nickname:
            self._log(f"  检测到昵称: {nickname}")

        # 检查数据库中是否已有该UID的账号
        accounts = self.db.get_accounts(game)
        for acc in accounts:
            if uid and acc.uid == uid:
                # 已有账号，直接返回
                self._log(f"  使用已有账号: {acc.nickname or acc.uid}")
                return acc

        # 对于明日方舟/终末地，如果没有UID，需要特殊处理
        if not uid and game in ["arknights", "endfield"]:
            if accounts:
                # 已有账号，询问用户是使用已有还是输入新UID
                reply = QMessageBox.question(
                    self, "选择账号",
                    f"检测到已有{GAME_NAMES.get(game, game)}账号：\n"
                    f"  {accounts[0].nickname or accounts[0].uid}\n\n"
                    "是否使用该账号？\n"
                    "（选择\"否\"可以输入新的账号UID）",
                    QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
                )
                if reply == QMessageBox.StandardButton.Yes:
                    self._log(f"  使用已有账号: {accounts[0].nickname or accounts[0].uid}")
                    return accounts[0]

            # 弹出输入对话框让用户手动输入UID
            uid, ok = QInputDialog.getText(
                self, "输入账号UID",
                f"请输入{GAME_NAMES.get(game, game)}的账号UID："
            )
            # 用户取消或输入为空
            if not ok or not uid.strip():
                return None
            uid = uid.strip()

        # 如果仍然没有UID，复用该游戏的第一个已有账号
        if not uid and accounts:
            self._log(f"  复用已有账号: {accounts[0].nickname or accounts[0].uid}")
            return accounts[0]

        # 创建新账号对象
        account = Account(
            game=game,
            uid=uid or "",
            nickname=nickname or "",
            server=region,
        )
        # 将账号写入数据库，获取自增ID
        account.id = self.db.add_account(account)
        self._log(f"  自动创建账号: {uid or '(待确认)'} ({region})")
        return account

    def _on_game_fetch_done(self, records):
        """单个游戏获取完成的槽函数（批量获取模式）

        处理获取到的记录：
            1. 检查并更新UID（从fetcher或记录中提取）
            2. 生成默认昵称
            3. 将原始记录转换为GachaRecord对象
            4. 写入数据库并计算保底数
            5. 继续获取下一个游戏

        参数：
            records (list): 获取到的原始记录列表（dict或GachaRecord混合）
        """
        try:
            if not records:
                self._log("  未获取到记录")
            else:
                account = self.main_window.get_current_account()
                if account:
                    # 从fetcher获取检测到的UID
                    detected_uid = getattr(self.fetch_thread, 'detected_uid', '')
                    # 如果fetcher检测到了不同的UID，更新账号
                    if detected_uid and detected_uid != account.uid:
                        self._log(f"  检测到UID: {detected_uid}")
                        account.uid = detected_uid

                    # 如果还是没有UID，尝试从记录中提取
                    if not account.uid:
                        from fetchers.mihoyo.api import MihoyoAPI
                        uid = MihoyoAPI.get_uid_from_records(records)
                        if uid:
                            account.uid = uid
                            self._log(f"  从记录中提取UID: {uid}")

                    # 生成默认昵称（如果有UID但还没有昵称）
                    if not account.nickname and account.uid:
                        account.nickname = self._generate_nickname(account.game, account.uid)

                    # 更新账号信息到数据库
                    self.db.update_account(account)
                    # 通知主窗口设置当前账号
                    self.main_window.set_account(account)

                    # 将原始记录（dict格式）转换为GachaRecord对象
                    gacha_records = []
                    for raw in records:
                        if isinstance(raw, dict):
                            # 字典格式的原始记录，通过MihoyoAPI解析
                            record = MihoyoAPI.parse_record(raw, account.game, account.id)
                            gacha_records.append(record)
                        else:
                            # 已经是GachaRecord对象，直接使用
                            gacha_records.append(raw)

                    # 写入数据库，返回新增记录数（去重后的）
                    new_count = self.db.add_records(gacha_records)
                    # 计算重复记录数
                    skipped_count = len(records) - new_count

                    if new_count > 0:
                        self._log(f"  成功导入 {new_count} 条新记录")
                        # 计算保底数（每条记录距上次出金的抽数）
                        self._log("  正在计算保底数...")
                        self.db.calculate_pity_counts(account.id)
                        self._log("  保底数计算完成")
                    if skipped_count > 0:
                        self._log(f"  跳过 {skipped_count} 条重复记录")
        except Exception as e:
            self._log(f"  ✗ 处理记录时出错: {e}")

        # 继续获取下一个游戏
        self._current_fetch_index += 1
        self._fetch_next_game()

    def _on_game_fetch_error(self, error_msg):
        """单个游戏获取失败的槽函数（批量获取模式）

        分析错误类型并给出相应提示：
            - 取消：停止所有获取
            - authkey过期：提示重新进入抽卡页面
            - 网络错误：提示检查网络连接

        参数：
            error_msg (str): 错误描述文字
        """
        self._log(f"  ✗ 获取失败: {error_msg}")
        # 如果是用户取消
        if "取消" in error_msg:
            self._set_fetching(False)
            self._log("已取消获取")
            return
        # 如果是authkey过期
        if "authkey" in error_msg.lower() or "expired" in error_msg.lower() or "过期" in error_msg:
            self._log("  提示: authkey已过期，请重新打开游戏进入抽卡记录页面")
        # 如果是网络错误
        elif "网络" in error_msg or "timeout" in error_msg.lower():
            self._log("  提示: 网络请求失败，请检查网络连接")
        # 跳过当前游戏，继续获取下一个
        self._current_fetch_index += 1
        self._fetch_next_game()

    def _url_fetch(self):
        """URL获取 - 从用户粘贴的URL获取抽卡记录

        流程：
            1. 获取输入框中的URL
            2. 使用URLParser自动识别游戏类型
            3. 对终末地/明日方舟的Token特殊处理（通过API获取UID）
            4. 自动检测/创建账号
            5. 切换到对应游戏
            6. 启动FetchThread获取记录

        URL/Token格式说明：
            - 原神/铁道/绝区零/鸣潮：完整的抽卡记录API URL
            - 终末地/明日方舟：鹰角账号Token（短token或长token）
        """
        url = self.url_input.text().strip()
        if not url:
            QMessageBox.warning(self, "提示", "请输入URL或Token")
            return

        # 使用URLParser自动检测游戏类型
        from fetchers.url_parser import URLParser
        parsed = URLParser.parse(url)
        game = parsed.get("game", "")

        # 终末地/明日方舟的特殊处理：
        # 如果输入不是HTTP URL，且当前游戏是endfield/arknights，
        # 则将输入当作账号Token处理
        if not game and not url.startswith("http"):
            current = self.main_window.get_current_game()
            if current in ["endfield", "arknights"]:
                game = current

        # 如果URLParser无法识别游戏，使用主窗口当前的游戏
        if not game:
            game = self.main_window.get_current_game()
            self._log(f"无法自动识别游戏，使用当前游戏: {GAME_NAMES.get(game, game)}")
        else:
            self._log(f"自动识别游戏: {GAME_NAMES.get(game, game)}")

        # 对于明日方舟和终末地，先通过Token获取UID
        uid = None
        if game in ["arknights", "endfield"] and not url.startswith("http"):
            uid = self._get_uid_from_token(game, url)
            if uid:
                self._log(f"  从Token检测到UID: {uid}")

        # 自动检测或创建账号
        account = self._auto_detect_account(game, url, uid)
        if not account:
            return

        # 切换到对应游戏和账号
        self.main_window._on_game_changed(game)
        self.main_window.set_account(account)

        # 设置UI为"获取中"状态
        self._set_fetching(True)
        self._log(f"开始从URL获取 {GAME_NAMES.get(game, game)} 抽卡记录...")

        # 获取最新记录时间，用于增量获取
        latest_time = None
        records = self.db.get_records(account.id)
        if records:
            latest_time = max(r.time for r in records if r.time)
            self._log(f"已有记录，从 {latest_time} 开始增量获取")

        # 创建并启动获取线程
        # 注意：单个获取模式的完成处理用 _on_fetch_done（不继续下一个游戏）
        self.fetch_thread = FetchThread(game, url=url, account_id=account.id, latest_time=latest_time)
        self.fetch_thread.progress.connect(self._on_progress)
        self.fetch_thread.finished.connect(self._on_fetch_done)
        self.fetch_thread.error.connect(self._on_fetch_error)
        self.fetch_thread.start()

    def _on_progress(self, message, progress):
        """进度更新的槽函数

        由FetchThread的progress信号触发，更新状态标签和进度条。

        参数：
            message (str): 进度描述文字
            progress (float): 进度值，0.0-1.0之间，-1表示不确定进度
        """
        # 更新状态标签
        self.status_label.setText(message)
        # 如果进度值有效（>0），更新进度条
        if progress > 0:
            self.progress_bar.setValue(int(progress * 100))
        # 追加到日志
        self._log(message)

    def _on_fetch_done(self, records):
        """获取完成的槽函数（单个获取模式）

        处理获取到的记录，逻辑与 _on_game_fetch_done 基本相同，
        但多了弹窗显示导入结果和刷新主窗口。

        参数：
            records (list): 获取到的原始记录列表
        """
        # 恢复UI状态
        self._set_fetching(False)

        if not records:
            self._log("未获取到任何记录")
            QMessageBox.information(self, "提示", "未获取到任何记录")
            return

        account = self.main_window.get_current_account()
        if account:
            # 获取并更新UID
            detected_uid = getattr(self.fetch_thread, 'detected_uid', '')
            if detected_uid and detected_uid != account.uid:
                self._log(f"检测到UID: {detected_uid}")
                account.uid = detected_uid

            # 从记录中提取UID（如果还没有）
            if not account.uid:
                from fetchers.mihoyo.api import MihoyoAPI
                uid = MihoyoAPI.get_uid_from_records(records)
                if uid:
                    account.uid = uid
                    self._log(f"从记录中提取UID: {uid}")

            # 生成默认昵称
            if not account.nickname and account.uid:
                account.nickname = self._generate_nickname(account.game, account.uid)

            # 更新账号信息
            self.db.update_account(account)
            self.main_window.set_account(account)

            # 将原始记录转换为GachaRecord对象
            gacha_records = []
            for raw in records:
                if isinstance(raw, dict):
                    record = MihoyoAPI.parse_record(raw, account.game, account.id)
                    gacha_records.append(record)
                else:
                    gacha_records.append(raw)

            # 写入数据库
            new_count = self.db.add_records(gacha_records)
            skipped_count = len(records) - new_count

            if new_count > 0:
                self._log(f"成功导入 {new_count} 条新记录")
                self._log("正在计算保底数...")
                self.db.calculate_pity_counts(account.id)
                self._log("保底数计算完成")
            if skipped_count > 0:
                self._log(f"跳过 {skipped_count} 条重复记录")

            # 弹窗显示导入结果摘要
            QMessageBox.information(
                self, "导入完成",
                f"新记录: {new_count} 条\n重复记录: {skipped_count} 条\n总计获取: {len(records)} 条"
            )
            # 刷新主窗口所有页面
            self.main_window.refresh_all()
        else:
            self._log("错误：未找到账号")

    def _on_fetch_error(self, error_msg):
        """获取失败的槽函数（单个获取模式）

        参数：
            error_msg (str): 错误描述文字
        """
        # 恢复UI状态
        self._set_fetching(False)
        # 如果是用户取消
        if "取消" in error_msg:
            self._log("已取消获取")
            return
        self._log(f"错误：{error_msg}")
        # 根据错误类型给出具体提示
        if "authkey" in error_msg.lower() or "expired" in error_msg.lower():
            self._log("提示: authkey已过期，请重新打开游戏进入抽卡记录页面")
            QMessageBox.critical(self, "获取失败", f"{error_msg}\n\nauthkey已过期，请重新打开游戏进入抽卡记录页面，然后切回本程序重试。")
        elif "网络" in error_msg or "timeout" in error_msg.lower():
            self._log("提示: 网络请求失败，请检查网络连接")
            QMessageBox.critical(self, "获取失败", f"{error_msg}\n\n请检查网络连接后重试。")
        else:
            QMessageBox.critical(self, "获取失败", error_msg)

    def _cancel_fetch(self):
        """取消当前获取

        通过以下步骤实现取消：
            1. 调用 FetchThread.cancel() 设置取消标志位
            2. 终止可能存在的代理子进程（某些fetcher会启动子进程）
            3. 更新取消按钮状态为"取消中..."
        """
        if self.fetch_thread and self.fetch_thread.isRunning():
            # 设置取消标志位
            self.fetch_thread.cancel()
            # 尝试终止代理子进程（如果有）
            try:
                # 获取fetcher实例
                fetcher = getattr(self.fetch_thread, '_fetcher_instance', None)
                if fetcher:
                    # 获取代理子进程引用
                    proc = getattr(fetcher, '_proxy_proc', None)
                    # 检查进程是否还在运行
                    if proc and proc.poll() is None:
                        # 强制终止子进程
                        proc.kill()
            except Exception:
                # 忽略终止过程中的异常
                pass
            # 更新取消按钮状态（防止重复点击）
            self.cancel_btn.setEnabled(False)
            self.cancel_btn.setText("取消中...")
            self._log("正在取消获取...")

    def _get_uid_from_token(self, game: str, token: str) -> str:
        """通过token获取UID（仅明日方舟和终末地）

        支持两种token类型：
            1. 短token（<50字符）：鹰角账号token，需要通过OAuth流程交换
               流程：hg_token → app_token → binding_list → UID
            2. 长token（>=50字符）：u8_token，无法直接获取UID

        API流程（短token）：
            1. POST https://as.hypergryph.com/user/oauth2/v2/grant
               body: {"type": 1, "appCode": "be36d44aa36bfb5b", "token": hg_token}
               返回: app_token
            2. GET https://binding-api-account-prod.hypergryph.com/account/binding/v1/binding_list
               params: {"token": app_token, "appCode": game}
               返回: 绑定的角色列表，包含UID

        参数：
            game (str): 游戏ID（"arknights"或"endfield"）
            token (str): 鹰角账号token

        返回：
            str|None: UID字符串，获取失败返回None
        """
        try:
            # 延迟导入requests库
            import requests as req

            self._log(f"  正在从Token获取UID... (Token长度: {len(token)})")

            # 短token是鹰角账号token，需要交换
            if len(token) < 50:
                self._log(f"  检测到鹰角账号Token，正在交换...")
                # 第一步：将hg_token交换为app_token
                grant_resp = req.post(
                    "https://as.hypergryph.com/user/oauth2/v2/grant",
                    json={"type": 1, "appCode": "be36d44aa36bfb5b", "token": token},
                    timeout=15,
                )
                grant_data = grant_resp.json()
                app_token = grant_data.get("data", {}).get("token")
                if not app_token:
                    self._log(f"  ✗ 获取app_token失败: {grant_data.get('msg', '未知错误')}")
                    return None

                # 第二步：通过app_token获取绑定的角色列表
                binding_resp = req.get(
                    "https://binding-api-account-prod.hypergryph.com/account/binding/v1/binding_list",
                    params={"token": app_token, "appCode": game},
                    timeout=15,
                )
                binding_data = binding_resp.json()
                # 从返回数据中提取应用列表
                apps = binding_data.get("data", {}).get("list", [])

                # 遍历应用列表，找到匹配的游戏
                for app in apps:
                    if app.get("appCode") == game:
                        # 遍历绑定的角色列表，提取UID
                        for binding in app.get("bindingList", []):
                            uid = binding.get("uid", "")
                            if uid:
                                self._log(f"  ✓ 获取到UID: {uid}")
                                return str(uid)
                self._log(f"  ✗ 未找到{game}的绑定角色")
            else:
                # 长token（u8_token）无法直接获取UID
                self._log(f"  Token是u8_token（长token），无法直接获取UID")
        except Exception as e:
            self._log(f"  ✗ 获取UID失败: {str(e)}")
        # 获取失败返回None
        return None

    def _set_fetching(self, fetching):
        """设置UI的获取状态

        在获取开始时禁用操作按钮、显示进度条和取消按钮；
        获取结束后恢复按钮、隐藏进度条和取消按钮。

        参数：
            fetching (bool): True表示正在获取中，False表示空闲状态
        """
        # 禁用/启用自动获取按钮
        self.auto_fetch_btn.setEnabled(not fetching)
        # 禁用/启用URL获取按钮
        self.url_fetch_btn.setEnabled(not fetching)
        # 显示/隐藏进度条
        self.progress_bar.setVisible(fetching)
        # 显示/隐藏取消按钮
        self.cancel_btn.setVisible(fetching)
        # 启用/禁用取消按钮
        self.cancel_btn.setEnabled(fetching)
        # 重置取消按钮文字
        self.cancel_btn.setText("取消获取")
        if fetching:
            # 设置进度条为不确定模式（显示来回滚动的动画）
            # range(0, 0) 表示不确定进度
            self.progress_bar.setRange(0, 0)
        else:
            # 恢复进度条为正常模式
            self.progress_bar.setRange(0, 100)
            self.progress_bar.setValue(0)

    def _update_login_btn_visibility(self):
        """根据选择的游戏更新登录按钮的显示状态

        登录获取仅支持终末地和明日方舟，因此：
            - 选择"全部游戏" → 显示登录按钮
            - 选择终末地/明日方舟 → 显示登录按钮
            - 选择其他游戏 → 隐藏登录按钮
        """
        selected_id = self.auto_game_combo.currentData()
        # 判断是否应该显示登录按钮
        show_login = selected_id in ["all", "endfield", "arknights"]
        self.login_btn.setVisible(show_login)

    def _paste_from_clipboard(self):
        """从系统剪贴板粘贴文本到URL输入框

        使用QApplication.clipboard()获取系统剪贴板，
        读取其中的文本内容并设置到URL输入框。
        """
        from PySide6.QtWidgets import QApplication
        # 获取系统剪贴板实例
        clipboard = QApplication.clipboard()
        # 将剪贴板文本设置到URL输入框
        self.url_input.setText(clipboard.text())

    def _login_fetch(self):
        """登录获取 - 根据当前游戏打开对应的登录窗口

        仅支持终末地和明日方舟。
        根据当前选中的游戏分发到对应的登录对话框。
        """
        game = self.main_window.get_current_game()

        if game == "endfield":
            # 终末地登录
            self._login_endfield()
        elif game == "arknights":
            # 明日方舟登录
            self._login_arknights()
        else:
            # 其他游戏不支持登录获取
            QMessageBox.information(self, "提示", "登录获取仅支持终末地和明日方舟。")

    def _login_endfield(self):
        """终末地登录 - API版

        打开终末地登录对话框（LoginApiDialog），
        用户登录鹰角账号后获取Token，
        自动填入URL输入框并触发URL获取。

        流程：
            1. 创建并显示 LoginApiDialog（模态对话框）
            2. 用户登录成功后获取 framework_token
            3. 将token填入URL输入框
            4. 延迟100ms后自动触发URL获取（QTimer.singleShot避免UI阻塞）
        """
        try:
            # 延迟导入登录对话框（避免循环引用）
            from ui.widgets.login_dialog_api import LoginApiDialog

            # 创建登录对话框（模态）
            dialog = LoginApiDialog(self)
            # 显示对话框并等待用户操作
            if dialog.exec() == QDialog.DialogCode.Accepted:
                # 用户登录成功，获取token
                token = dialog.get_framework_token()
                if token:
                    # 将token填入URL输入框
                    self.url_input.setText(token)
                    self._log(f"✓ 终末地登录成功，Token: {token[:20]}...")
                    # 延迟100ms后触发URL获取（给UI时间更新）
                    QTimer.singleShot(100, self._url_fetch)
                else:
                    self._log("✗ 未获取到凭证")
                    QMessageBox.warning(self, "提示", "未获取到凭证")
            else:
                self._log("登录已取消")
        except Exception as e:
            self._log(f"✗ 登录出错: {type(e).__name__}: {str(e)}")

    def _login_arknights(self):
        """明日方舟登录 - API版

        打开明日方舟登录对话框（ArknightsLoginApiDialog），
        用户登录鹰角账号后获取Token，
        自动填入URL输入框并触发URL获取。

        流程与 _login_endfield 类似，使用不同的对话框类。
        """
        try:
            # 延迟导入登录对话框
            from ui.widgets.arknights_login_api import ArknightsLoginApiDialog

            dialog = ArknightsLoginApiDialog(self)
            if dialog.exec() == QDialog.DialogCode.Accepted:
                # 获取token
                token = dialog.get_token()
                if token:
                    self.url_input.setText(token)
                    self._log(f"✓ 明日方舟登录成功，Token: {token[:20]}...")
                    # 延迟触发URL获取
                    QTimer.singleShot(100, self._url_fetch)
                else:
                    self._log("✗ 未获取到 Token")
                    QMessageBox.warning(self, "提示", "未获取到 Token，请重试。")
            else:
                self._log("登录已取消")
        except Exception as e:
            self._log(f"✗ 登录出错: {type(e).__name__}: {str(e)}")

    def _import_file(self, file_type):
        """文件导入

        打开文件选择对话框，让用户选择要导入的文件，
        然后解析文件内容并写入数据库。

        参数：
            file_type (str): 文件类型标识，"json"、"excel"或"csv"

        流程：
            1. 弹出文件选择对话框（过滤器根据file_type显示对应格式）
            2. 获取当前游戏和账号
            3. 如果没有账号，弹出创建账号对话框
            4. 调用 _parse_file() 解析文件
            5. 将解析到的记录写入数据库
        """
        # 文件类型到过滤器字符串的映射
        filters = {
            "json": "JSON 文件 (*.json)",
            "excel": "Excel 文件 (*.xlsx *.xls)",
            "csv": "CSV 文件 (*.csv)",
        }
        # 弹出文件选择对话框
        filepath, _ = QFileDialog.getOpenFileName(
            self, "选择文件", "", filters.get(file_type, "")
        )
        # 用户取消选择
        if not filepath:
            return

        # 获取当前游戏和账号
        game = self.main_window.get_current_game()
        account = self.main_window.get_current_account()

        # 如果没有当前账号，弹出创建账号对话框
        if not account:
            account = self._ensure_account(game)
            if not account:
                return

        try:
            # 解析文件内容
            records = self._parse_file(filepath, file_type, game, account.id)
            if records:
                # 写入数据库
                count = self.db.add_records(records)
                self._log(f"从文件导入 {count} 条新记录")
                QMessageBox.information(self, "导入成功", f"成功导入 {count} 条新记录！")
                # 刷新主窗口所有页面
                self.main_window.refresh_all()
            else:
                QMessageBox.warning(self, "提示", "文件中没有找到有效记录")
        except Exception as e:
            self._log(f"导入失败：{str(e)}")
            QMessageBox.critical(self, "导入失败", f"解析文件失败：\n{str(e)}")

    def _parse_file(self, filepath, file_type, game, account_id):
        """解析导入文件

        根据文件类型解析不同格式的抽卡记录文件，返回GachaRecord列表。

        支持的格式：
            JSON:
                - 小黑盒格式: {"info": {...}, "data": {"timestamp": {"c": [...], "p": "..."}, ...}}
                - UIGF格式: {"info": {...}, "list": [{记录}, ...]}
                - 通用格式: {"list": [...]} 或 {"records": [...]} 或 直接数组
            CSV:
                - 表头行: name, rarity, is_featured, time, ...
                - 每行一条记录
            Excel:
                - 表头行: name, rarity, is_featured, time, ...
                - 使用openpyxl读取

        参数：
            filepath (str): 文件绝对路径
            file_type (str): 文件类型，"json"/"excel"/"csv"
            game (str): 游戏ID（用于默认值和格式判断）
            account_id (int): 账号ID（用于关联记录）

        返回：
            list[GachaRecord]: 解析到的抽卡记录列表

        异常：
            - JSON解析失败会抛出 json.JSONDecodeError
            - Excel文件需要openpyxl库，未安装会抛出RuntimeError
            - CSV编码使用utf-8-sig（兼容BOM头）
        """
        records = []

        if file_type == "json":
            # 以UTF-8编码读取JSON文件
            with open(filepath, "r", encoding="utf-8") as f:
                data = json.load(f)

            # ===== 自动检测JSON格式 =====
            # 小黑盒格式: {"info": {...}, "data": {"timestamp": {"c": [...], "p": "..."}, ...}}
            if isinstance(data, dict) and "info" in data and "data" in data and isinstance(data["data"], dict):
                records = self._parse_xiaoheihe(data, game, account_id)
            # UIGF格式: {"info": {...}, "list": [{记录}, ...]}
            elif isinstance(data, dict) and "info" in data and "list" in data:
                records = self._parse_uigf(data, game, account_id)
            else:
                # ===== 通用格式兼容 =====
                # 尝试从多种常见字段名中提取记录列表
                if isinstance(data, dict):
                    if "list" in data:
                        data = data["list"]
                    elif "records" in data:
                        data = data["records"]
                    else:
                        data = data.get("list", data.get("records", data.get("data", [])))

                # 遍历每条记录，构造GachaRecord对象
                for item in data:
                    records.append(GachaRecord(
                        account_id=account_id,
                        game=item.get("game", game),
                        # 兼容pool_type和gacha_type两种字段名
                        pool_type=item.get("pool_type", item.get("gacha_type", "character")),
                        # 兼容item_name和name两种字段名
                        item_name=item.get("item_name", item.get("name", "未知")),
                        # 兼容item_type和type两种字段名
                        item_type=item.get("item_type", item.get("type", "")),
                        # 兼容rarity和rank_type两种字段名，默认3星
                        rarity=int(item.get("rarity", item.get("rank_type", 3))),
                        # 兼容is_featured和is_up两种字段名，默认非UP
                        is_featured=bool(item.get("is_featured", item.get("is_up", False))),
                        time=item.get("time", ""),
                        pity_count=int(item.get("pity_count", 0)),
                    ))

        elif file_type == "csv":
            # 使用utf-8-sig编码读取CSV（兼容带BOM头的文件）
            with open(filepath, "r", encoding="utf-8-sig") as f:
                # DictReader自动将每行解析为字典（以第一行为键名）
                reader = csv.DictReader(f)
                for row in reader:
                    records.append(GachaRecord(
                        account_id=account_id,
                        game=row.get("game", game),
                        pool_type=row.get("pool_type", "character"),
                        item_name=row.get("item_name", row.get("name", "未知")),
                        item_type=row.get("item_type", ""),
                        rarity=int(row.get("rarity", 3)),
                        # "is_featured"字段的多种真值判断
                        is_featured=row.get("is_featured", "").lower() in ("true", "1", "是"),
                        time=row.get("time", ""),
                    ))

        elif file_type == "excel":
            try:
                # 使用openpyxl读取Excel文件
                import openpyxl
                # read_only=True模式更快，适合只读场景
                wb = openpyxl.load_workbook(filepath, read_only=True)
                ws = wb.active
                # 读取第一行作为表头
                headers = [cell.value for cell in next(ws.iter_rows(max_row=1))]
                # 从第二行开始读取数据行
                for row in ws.iter_rows(min_row=2, values_only=True):
                    # 将行数据与表头合并为字典
                    item = dict(zip(headers, row))
                    records.append(GachaRecord(
                        account_id=account_id,
                        game=str(item.get("game", game)),
                        pool_type=str(item.get("pool_type", "character")),
                        item_name=str(item.get("item_name", item.get("name", "未知"))),
                        item_type=str(item.get("item_type", "")),
                        rarity=int(item.get("rarity", 3)),
                        is_featured=bool(item.get("is_featured", False)),
                        time=str(item.get("time", "")),
                    ))
            except ImportError:
                # openpyxl未安装时抛出明确的错误提示
                raise RuntimeError("需要安装 openpyxl 才能导入 Excel 文件")

        return records

    def _parse_xiaoheihe(self, data: dict, game: str, account_id: int) -> list:
        """解析小黑盒导出格式的JSON

        小黑盒（Xiaoheihe）是一款游戏数据查询工具，其导出格式：
        {
            "info": {"uid": "123456789", ...},
            "data": {
                "1699000000": {  // Unix时间戳（秒）
                    "c": [       // 角色/物品列表
                        ["角色名", 5, true],  // [名称, 星级(0-5), 是否UP]
                    ],
                    "p": "限定寻访-XXX"  // 卡池名
                },
                ...
            }
        }

        参数：
            data (dict): 解析后的JSON数据
            game (str): 游戏ID
            account_id (int): 账号ID

        返回：
            list[GachaRecord]: 解析到的抽卡记录列表

        特殊处理：
            - 小黑盒的星级范围是0-5，明日方舟需要+1转为1-6
            - 时间戳字符串需要转换为datetime格式
            - 生成唯一item_id以避免与API获取的记录重复
        """
        from datetime import datetime
        from core.models import get_max_rarity

        records = []
        # 从info中获取UID
        uid = str(data.get("info", {}).get("uid", ""))

        # 小黑盒的星级和API一致（0-5），明日方舟的星级是1-6，需要+1转换
        rarity_offset = 1 if game == "arknights" else 0

        # 遍历data中的每个时间戳条目
        for ts_str, entry in data.get("data", {}).items():
            # 解析时间戳字符串为datetime格式
            try:
                ts = int(ts_str)
                # 将Unix时间戳转换为"YYYY-MM-DD HH:MM:SS"格式
                time_str = datetime.fromtimestamp(ts).strftime("%Y-%m-%d %H:%M:%S")
            except (ValueError, OSError):
                # 时间戳解析失败，使用空字符串
                time_str = ""

            # 获取卡池名
            pool_name = entry.get("p", "")
            # 获取角色/物品列表
            chars = entry.get("c", [])

            # 根据游戏确定pool_type（卡池类型分组）
            if game == "arknights":
                # 明日方舟：根据卡池名判断pool_type
                pool_type = self._get_arknights_pool_type(pool_name)
            else:
                # 其他游戏默认为角色池
                pool_type = "character"

            # 遍历每个抽卡记录
            for idx, char in enumerate(chars):
                if len(char) < 2:
                    # 至少需要2个元素（名称和星级）
                    continue
                char_name = char[0]
                # 星级加上偏移量（明日方舟+1）
                rarity = int(char[1]) + rarity_offset
                # 是否UP（第3个元素，可选）
                is_featured = bool(char[2]) if len(char) > 2 else False

                # 生成唯一item_id：角色名_时间（与游戏API获取格式一致，避免重复）
                item_id = f"{char_name}_{time_str}"

                records.append(GachaRecord(
                    account_id=account_id,
                    game=game,
                    pool_type=pool_type,
                    pool_name=pool_name,
                    item_id=item_id,
                    item_name=char_name,
                    item_type="CHAR",
                    rarity=rarity,
                    is_featured=is_featured,
                    count=1,
                    time=time_str,
                ))

        return records

    def _parse_uigf(self, data: dict, game: str, account_id: int) -> list:
        """解析 UIGF (Unified Interchangeable GachaLog Format) 标准格式

        UIGF 是跨游戏通用的抽卡记录交换格式，格式如下：
        {
            "info": {
                "export_timestamp": 1699000000,
                "export_app": "...",
                ...
            },
            "list": [
                {
                    "uid": "123456789",
                    "gacha_type": "301",      // 卡池类型编号（游戏特定）
                    "time": "2023-11-01 12:00:00",
                    "name": "纳西妲",
                    "item_type": "角色",
                    "rank_type": "5",
                    "id": "...",
                    ...
                },
                ...
            ]
        }

        参数：
            data (dict): 解析后的JSON数据
            game (str): 游戏ID
            account_id (int): 账号ID

        返回：
            list[GachaRecord]: 解析到的抽卡记录列表

        内部逻辑：
            1. 根据游戏选择对应的 gacha_type → pool_type 映射表
            2. 遍历list中的每条记录，转换pool_type
            3. 将UIGF格式的字段映射到GachaRecord的字段
            4. 调用MihoyoAPI.parse_record()进行标准化处理
        """
        from fetchers.mihoyo.api import MihoyoAPI

        # UIGF gacha_type 到 pool_type 的映射表
        # 不同游戏的gacha_type编号不同，需要分别映射
        UIGF_TYPE_MAP = {
            "genshin": {
                "100": "beginner",       # 新手祈愿
                "200": "standard",       # 常驻祈愿
                "301": "character",      # 角色活动祈愿
                "302": "weapon",         # 武器活动祈愿
                "400": "character",      # 角色活动祈愿（旧格式）
                "500": "chronicled",     # 集录祈愿
            },
            "starrail": {
                "1": "standard",         # 常跃迁
                "2": "beginner",         # 新手跃迁
                "11": "character",       # 角色活动跃迁
                "12": "weapon",          # 光锥活动跃迁
                "13": "collab",          # 联动跃迁
                "14": "collab_weapon",   # 联动光锥跃迁
            },
            "zzz": {
                "1001": "standard",       # 常驻频段
                "2001": "character",      # 角色频段
                "3001": "weapon",         # 武器频段
                "4001": "special",        # 特殊角色频段
                "5001": "special_weapon", # 特殊武器频段
                "6001": "bangboo",        # 邦布频段
                # 兼容 API 新短格式
                "1": "standard",
                "2": "character",
                "3": "weapon",
                "4": "special",
                "5": "special_weapon",
                "6": "bangboo",
            },
        }

        # 获取当前游戏的映射表（如果游戏不在映射表中，使用空映射）
        type_map = UIGF_TYPE_MAP.get(game, {})
        records = []

        # 遍历UIGF格式中的记录列表
        for item in data.get("list", []):
            # 获取gacha_type，兼容uigf_gacha_type字段
            gacha_type = str(item.get("gacha_type", item.get("uigf_gacha_type", "")))
            # 通过映射表将gacha_type转为pool_type
            pool_type = type_map.get(gacha_type, "character")

            # 注入_pool_type字段供MihoyoAPI.parse_record使用
            item["_pool_type"] = pool_type
            # 使用MihoyoAPI标准化解析记录
            record = MihoyoAPI.parse_record(item, game, account_id)
            records.append(record)

        return records

    def _get_arknights_pool_type(self, pool_name: str) -> str:
        """根据明日方舟卡池名称返回保底分组类型

        明日方舟的卡池名称可能包含特定关键词，用于判断该卡池属于哪种保底机制：
            - "限定"/"联动"等 → limited（独立寻访，不共享保底）
            - "中坚" → kernel（中坚寻访）
            - "标准"/"常驻"等 → standard（标准寻访）
            - 未识别 → limited（默认为独立寻访，因为大多数特定角色卡池是限定池）

        参数：
            pool_name (str): 卡池完整名称

        返回：
            str: 保底分组类型，"limited"/"kernel"/"standard"

        匹配优先级：
            1. 精确匹配（ARKNIGHTS_POOL_MECHANIC_MAP）
            2. 关键词匹配（遍历关键词列表）
            3. 默认返回"limited"
        """
        from core.models import ARKNIGHTS_POOL_MECHANIC_MAP, ARKNIGHTS_MECHANIC_TO_GROUP

        # 优先精确匹配
        mechanic = ARKNIGHTS_POOL_MECHANIC_MAP.get(pool_name, "")
        if mechanic:
            return ARKNIGHTS_MECHANIC_TO_GROUP.get(mechanic, "standard")

        # 关键词匹配
        # 限定池关键词
        limited_keywords = ["限定", "联动", "跨年", "归航", "启程", "承诺"]
        # 中坚池关键词
        kernel_keywords = ["中坚"]
        # 标准池关键词
        standard_keywords = ["标准", "常驻", "定向", "甄选"]

        # 按优先级检查关键词
        for kw in limited_keywords:
            if kw in pool_name:
                return "limited"
        for kw in kernel_keywords:
            if kw in pool_name:
                return "kernel"
        for kw in standard_keywords:
            if kw in pool_name:
                return "standard"

        # 未识别的卡池默认为独立寻访（limited）
        # 因为大多数特定角色卡池都是限定池
        return "limited"

    def _ensure_account(self, game):
        """确保有可用账号，没有则创建

        弹出两个输入对话框：第一个输入UID，第二个输入昵称（可选）。
        创建Account对象并写入数据库。

        参数：
            game (str): 游戏ID

        返回：
            Account|None: 创建的账号对象，用户取消返回None
        """
        from PySide6.QtWidgets import QInputDialog
        # 弹出UID输入对话框
        uid, ok = QInputDialog.getText(self, "创建账号", "请输入游戏UID:")
        # 用户取消或输入为空
        if not ok or not uid.strip():
            return None

        # 弹出昵称输入对话框（可选）
        nickname, _ = QInputDialog.getText(self, "设置昵称", "请输入昵称（可选）:")

        # 创建账号对象
        account = Account(
            game=game,
            uid=uid.strip(),
            nickname=nickname.strip() if nickname else "",
            server="cn",
        )
        # 写入数据库
        account.id = self.db.add_account(account)
        # 通知主窗口设置当前账号
        self.main_window.set_account(account)
        return account

    def refresh(self):
        """刷新页面状态

        同步左侧游戏列表的显示状态和顺序到自动获取的下拉框。
        当左侧游戏列表发生变化时（显示/隐藏/排序），调用此方法同步。
        """
        # 从主窗口获取游戏列表的可见性和排序信息
        visible = getattr(self.main_window, '_visible_games', [])
        order = getattr(self.main_window, '_game_order', [])
        # 完整游戏列表（ID → 名称）
        all_games = [
            ("genshin", "原神"), ("starrail", "星穹铁道"), ("zzz", "绝区零"),
            ("wutheringwaves", "鸣潮"), ("endfield", "终末地"), ("arknights", "明日方舟"),
        ]

        # 按左侧顺序筛选可见游戏
        ordered_visible = [g for g in order if g in visible]
        game_map = dict(all_games)

        # 阻塞信号，避免clear()触发不必要的信号
        self.auto_game_combo.blockSignals(True)
        self.auto_game_combo.clear()
        # 重建游戏选项列表
        self._auto_game_options = [("all", "全部游戏")]
        self.auto_game_combo.addItem("全部游戏", "all")
        for gid in ordered_visible:
            gname = game_map.get(gid, gid)
            self._auto_game_options.append((gid, gname))
            self.auto_game_combo.addItem(gname, gid)
        # 恢复信号
        self.auto_game_combo.blockSignals(False)

        # 同步当前游戏选择（高亮显示当前游戏）
        current_game = self.main_window.get_current_game()
        for i, (gid, _) in enumerate(self._auto_game_options):
            if gid == current_game:
                self.auto_game_combo.setCurrentIndex(i)
                break
