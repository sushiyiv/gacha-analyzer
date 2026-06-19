"""文件解析模块 - 从 import_widget 提取的纯解析逻辑"""

import json
import csv
from core.models import GachaRecord


# UIGF gacha_type 到 pool_type 的映射
UIGF_TYPE_MAP = {
    "genshin": {
        "100": "beginner",
        "200": "standard",
        "301": "character",
        "302": "weapon",
        "400": "character",
        "500": "chronicled",
    },
    "starrail": {
        "1": "standard",
        "2": "beginner",
        "11": "character",
        "12": "weapon",
        "13": "collab",
        "14": "collab_weapon",
    },
    "zzz": {
        "1001": "standard",
        "2001": "character",
        "3001": "weapon",
        "4001": "special",
        "5001": "special_weapon",
        "6001": "bangboo",
        "1": "standard",
        "2": "character",
        "3": "weapon",
        "4": "special",
        "5": "special_weapon",
        "6": "bangboo",
    },
}


def parse_file(filepath: str, file_type: str, game: str, account_id: int) -> list:
    """解析导入文件，返回 GachaRecord 列表"""
    if file_type == "json":
        return _parse_json(filepath, game, account_id)
    elif file_type == "csv":
        return _parse_csv(filepath, game, account_id)
    elif file_type == "excel":
        return _parse_excel(filepath, game, account_id)
    return []


def _parse_json(filepath: str, game: str, account_id: int) -> list:
    """解析 JSON 文件（自动检测格式）"""
    with open(filepath, "r", encoding="utf-8") as f:
        data = json.load(f)

    # 小黑盒格式: {"info": {...}, "data": {"id": {"c": [...], "p": "..."}, ...}}
    if isinstance(data, dict) and "info" in data and "data" in data and isinstance(data["data"], dict):
        return parse_xiaoheihe(data, game, account_id)
    # UIGF 格式: {"info": {...}, "list": [...]}
    elif isinstance(data, dict) and "info" in data and "list" in data:
        return parse_uigf(data, game, account_id)
    else:
        return _parse_generic_json(data, game, account_id)


def _parse_generic_json(data, game: str, account_id: int) -> list:
    """解析通用 JSON 格式"""
    if isinstance(data, dict):
        if "list" in data:
            data = data["list"]
        elif "records" in data:
            data = data["records"]
        else:
            data = data.get("list", data.get("records", data.get("data", [])))

    records = []
    for item in data:
        records.append(GachaRecord(
            account_id=account_id,
            game=item.get("game", game),
            pool_type=item.get("pool_type", item.get("gacha_type", "character")),
            item_name=item.get("item_name", item.get("name", "未知")),
            item_type=item.get("item_type", item.get("type", "")),
            rarity=int(item.get("rarity", item.get("rank_type", 3))),
            is_featured=bool(item.get("is_featured", item.get("is_up", False))),
            time=item.get("time", ""),
            pity_count=int(item.get("pity_count", 0)),
        ))
    return records


def parse_xiaoheihe(data: dict, game: str, account_id: int) -> list:
    """解析小黑盒导出格式

    格式: {"info": {...}, "data": {"timestamp": {"c": [[name, rarity, is_featured], ...], "p": "卡池名"}, ...}}
    """
    from datetime import datetime

    records = []

    rarity_offset = 1 if game == "arknights" else 0

    for ts_str, entry in data.get("data", {}).items():
        try:
            ts = int(ts_str)
            time_str = datetime.fromtimestamp(ts).strftime("%Y-%m-%d %H:%M:%S")
        except (ValueError, OSError):
            time_str = ""

        pool_name = entry.get("p", "")
        chars = entry.get("c", [])

        if game == "arknights":
            pool_type = get_arknights_pool_type(pool_name)
        else:
            pool_type = "character"

        for idx, char in enumerate(chars):
            if len(char) < 2:
                continue
            char_name = char[0]
            rarity = int(char[1]) + rarity_offset
            is_featured = bool(char[2]) if len(char) > 2 else False

            item_id = f"{char_name}_{time_str}"

            records.append(GachaRecord(
                account_id=account_id,
                game=game,
                pool_type=pool_type,
                pool_name=pool_name,
                item_id=item_id,
                item_name=char_name,
                item_type="CHAR",
                rarity=rarity,
                is_featured=is_featured,
                count=1,
                time=time_str,
            ))

    return records


def parse_uigf(data: dict, game: str, account_id: int) -> list:
    """解析 UIGF 标准格式"""
    from fetchers.mihoyo.api import MihoyoAPI

    type_map = UIGF_TYPE_MAP.get(game, {})
    records = []

    for item in data.get("list", []):
        gacha_type = str(item.get("gacha_type", item.get("uigf_gacha_type", "")))
        pool_type = type_map.get(gacha_type, "character")

        item["_pool_type"] = pool_type
        record = MihoyoAPI.parse_record(item, game, account_id)
        records.append(record)

    return records


def _parse_csv(filepath: str, game: str, account_id: int) -> list:
    """解析 CSV 文件"""
    records = []
    with open(filepath, "r", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for row in reader:
            records.append(GachaRecord(
                account_id=account_id,
                game=row.get("game", game),
                pool_type=row.get("pool_type", "character"),
                item_name=row.get("item_name", row.get("name", "未知")),
                item_type=row.get("item_type", ""),
                rarity=int(row.get("rarity", 3)),
                is_featured=row.get("is_featured", "").lower() in ("true", "1", "是"),
                time=row.get("time", ""),
            ))
    return records


def _parse_excel(filepath: str, game: str, account_id: int) -> list:
    """解析 Excel 文件"""
    try:
        import openpyxl
    except ImportError:
        raise RuntimeError("需要安装 openpyxl 才能导入 Excel 文件")

    records = []
    wb = openpyxl.load_workbook(filepath, read_only=True)
    ws = wb.active
    headers = [cell.value for cell in next(ws.iter_rows(max_row=1))]
    for row in ws.iter_rows(min_row=2, values_only=True):
        item = dict(zip(headers, row))
        records.append(GachaRecord(
            account_id=account_id,
            game=str(item.get("game", game)),
            pool_type=str(item.get("pool_type", "character")),
            item_name=str(item.get("item_name", item.get("name", "未知"))),
            item_type=str(item.get("item_type", "")),
            rarity=int(item.get("rarity", 3)),
            is_featured=bool(item.get("is_featured", False)),
            time=str(item.get("time", "")),
        ))
    return records


def get_arknights_pool_type(pool_name: str) -> str:
    """根据明日方舟卡池名返回保底分组"""
    from core.models import ARKNIGHTS_POOL_MECHANIC_MAP, ARKNIGHTS_MECHANIC_TO_GROUP

    mechanic = ARKNIGHTS_POOL_MECHANIC_MAP.get(pool_name, "")
    if mechanic:
        return ARKNIGHTS_MECHANIC_TO_GROUP.get(mechanic, "standard")

    limited_keywords = ["限定", "联动", "跨年", "归航", "启程", "承诺"]
    kernel_keywords = ["中坚"]
    standard_keywords = ["标准", "常驻", "定向", "甄选"]

    for kw in limited_keywords:
        if kw in pool_name:
            return "limited"
    for kw in kernel_keywords:
        if kw in pool_name:
            return "kernel"
    for kw in standard_keywords:
        if kw in pool_name:
            return "standard"

    return "limited"
